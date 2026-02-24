"""
Router para Reportes y Exportación a Excel.
Resúmenes diarios, mensuales y anuales + exportación Excel.
"""
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from datetime import date
from typing import Optional, List
import io

from app.database import get_db
from app.models.income import Income
from app.models.expense import Expense, ExpenseCategory
from app.schemas.reports import (
    DailySummary,
    MonthlySummary,
    YearlySummary,
    ExpenseByCategoryDetail,
)

router = APIRouter(prefix="/reportes", tags=["Reportes"])


def _gastos_por_categoria(db: Session, filtros) -> List[ExpenseByCategoryDetail]:
    """Obtener gastos agrupados por categoría con filtros aplicados."""
    query = db.query(
        Expense.categoria,
        func.sum(Expense.monto).label("total"),
        func.count(Expense.id).label("cantidad"),
    )
    for filtro in filtros:
        query = query.filter(filtro)
    resultados = query.group_by(Expense.categoria).all()

    return [
        ExpenseByCategoryDetail(
            categoria=r.categoria.value if hasattr(r.categoria, "value") else r.categoria,
            total=round(r.total, 2),
            cantidad=r.cantidad,
        )
        for r in resultados
    ]


# ─── RESUMEN DIARIO ────────────────────────────────────────────────
@router.get("/diario", response_model=DailySummary)
def resumen_diario(
    fecha: date = Query(..., description="Fecha del resumen (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
):
    """Obtener resumen de ingresos y gastos de un día específico."""
    # Ingresos del día
    ingresos = db.query(
        func.coalesce(func.sum(Income.monto), 0).label("total"),
        func.count(Income.id).label("cantidad"),
    ).filter(Income.fecha == fecha).first()

    # Gastos del día
    gastos = db.query(
        func.coalesce(func.sum(Expense.monto), 0).label("total"),
        func.count(Expense.id).label("cantidad"),
    ).filter(Expense.fecha == fecha).first()

    total_ingresos = round(float(ingresos.total), 2)
    total_gastos = round(float(gastos.total), 2)

    return DailySummary(
        fecha=fecha,
        total_ingresos=total_ingresos,
        total_gastos=total_gastos,
        balance=round(total_ingresos - total_gastos, 2),
        cantidad_ingresos=ingresos.cantidad,
        cantidad_gastos=gastos.cantidad,
        gastos_por_categoria=_gastos_por_categoria(db, [Expense.fecha == fecha]),
    )


# ─── RESUMEN MENSUAL ───────────────────────────────────────────────
@router.get("/mensual", response_model=MonthlySummary)
def resumen_mensual(
    anio: int = Query(..., description="Año"),
    mes: int = Query(..., ge=1, le=12, description="Mes (1-12)"),
    db: Session = Depends(get_db),
):
    """Obtener resumen de ingresos y gastos de un mes específico."""
    filtro_ing = [
        extract("year", Income.fecha) == anio,
        extract("month", Income.fecha) == mes,
    ]
    filtro_gas = [
        extract("year", Expense.fecha) == anio,
        extract("month", Expense.fecha) == mes,
    ]

    ingresos = db.query(
        func.coalesce(func.sum(Income.monto), 0).label("total"),
        func.count(Income.id).label("cantidad"),
    ).filter(*filtro_ing).first()

    gastos = db.query(
        func.coalesce(func.sum(Expense.monto), 0).label("total"),
        func.count(Expense.id).label("cantidad"),
    ).filter(*filtro_gas).first()

    # Contar días únicos con registros
    dias_ingresos = (
        db.query(func.distinct(Income.fecha))
        .filter(*filtro_ing)
        .count()
    )
    dias_gastos = (
        db.query(func.distinct(Expense.fecha))
        .filter(*filtro_gas)
        .count()
    )
    dias_con_registros = max(dias_ingresos, dias_gastos)

    total_ingresos = round(float(ingresos.total), 2)
    total_gastos = round(float(gastos.total), 2)

    return MonthlySummary(
        anio=anio,
        mes=mes,
        total_ingresos=total_ingresos,
        total_gastos=total_gastos,
        balance=round(total_ingresos - total_gastos, 2),
        cantidad_ingresos=ingresos.cantidad,
        cantidad_gastos=gastos.cantidad,
        gastos_por_categoria=_gastos_por_categoria(db, filtro_gas),
        dias_con_registros=dias_con_registros,
    )


# ─── RESUMEN ANUAL ─────────────────────────────────────────────────
@router.get("/anual", response_model=YearlySummary)
def resumen_anual(
    anio: int = Query(..., description="Año"),
    incluir_meses: bool = Query(False, description="Incluir desglose mensual"),
    db: Session = Depends(get_db),
):
    """Obtener resumen de ingresos y gastos de un año específico."""
    filtro_ing = [extract("year", Income.fecha) == anio]
    filtro_gas = [extract("year", Expense.fecha) == anio]

    ingresos = db.query(
        func.coalesce(func.sum(Income.monto), 0).label("total"),
        func.count(Income.id).label("cantidad"),
    ).filter(*filtro_ing).first()

    gastos = db.query(
        func.coalesce(func.sum(Expense.monto), 0).label("total"),
        func.count(Expense.id).label("cantidad"),
    ).filter(*filtro_gas).first()

    total_ingresos = round(float(ingresos.total), 2)
    total_gastos = round(float(gastos.total), 2)

    resumen_mensual = None
    if incluir_meses:
        resumen_mensual = []
        for m in range(1, 13):
            res = resumen_mensual_interno(anio, m, db)
            if res.cantidad_ingresos > 0 or res.cantidad_gastos > 0:
                resumen_mensual.append(res)

    return YearlySummary(
        anio=anio,
        total_ingresos=total_ingresos,
        total_gastos=total_gastos,
        balance=round(total_ingresos - total_gastos, 2),
        cantidad_ingresos=ingresos.cantidad,
        cantidad_gastos=gastos.cantidad,
        gastos_por_categoria=_gastos_por_categoria(db, filtro_gas),
        resumen_mensual=resumen_mensual,
    )


def resumen_mensual_interno(anio: int, mes: int, db: Session) -> MonthlySummary:
    """Función interna para generar resumen mensual (reutilizable)."""
    filtro_ing = [
        extract("year", Income.fecha) == anio,
        extract("month", Income.fecha) == mes,
    ]
    filtro_gas = [
        extract("year", Expense.fecha) == anio,
        extract("month", Expense.fecha) == mes,
    ]

    ingresos = db.query(
        func.coalesce(func.sum(Income.monto), 0).label("total"),
        func.count(Income.id).label("cantidad"),
    ).filter(*filtro_ing).first()

    gastos = db.query(
        func.coalesce(func.sum(Expense.monto), 0).label("total"),
        func.count(Expense.id).label("cantidad"),
    ).filter(*filtro_gas).first()

    dias_ingresos = db.query(func.distinct(Income.fecha)).filter(*filtro_ing).count()
    dias_gastos = db.query(func.distinct(Expense.fecha)).filter(*filtro_gas).count()

    total_ingresos = round(float(ingresos.total), 2)
    total_gastos = round(float(gastos.total), 2)

    return MonthlySummary(
        anio=anio,
        mes=mes,
        total_ingresos=total_ingresos,
        total_gastos=total_gastos,
        balance=round(total_ingresos - total_gastos, 2),
        cantidad_ingresos=ingresos.cantidad,
        cantidad_gastos=gastos.cantidad,
        gastos_por_categoria=_gastos_por_categoria(db, filtro_gas),
        dias_con_registros=max(dias_ingresos, dias_gastos),
    )


# ─── HISTORIAL ──────────────────────────────────────────────────────
@router.get("/historial/fechas-disponibles")
def fechas_disponibles(db: Session = Depends(get_db)):
    """Obtener los años y meses que tienen registros."""
    # Años con ingresos
    anios_ing = db.query(
        extract("year", Income.fecha).label("anio")
    ).distinct().all()

    # Años con gastos
    anios_gas = db.query(
        extract("year", Expense.fecha).label("anio")
    ).distinct().all()

    anios = sorted(set(
        [int(a.anio) for a in anios_ing] + [int(a.anio) for a in anios_gas]
    ), reverse=True)

    resultado = []
    for anio in anios:
        meses_ing = db.query(
            extract("month", Income.fecha).label("mes")
        ).filter(extract("year", Income.fecha) == anio).distinct().all()

        meses_gas = db.query(
            extract("month", Expense.fecha).label("mes")
        ).filter(extract("year", Expense.fecha) == anio).distinct().all()

        meses = sorted(set(
            [int(m.mes) for m in meses_ing] + [int(m.mes) for m in meses_gas]
        ))

        resultado.append({"anio": anio, "meses": meses})

    return resultado


# ─── EXPORTAR A EXCEL ───────────────────────────────────────────────
@router.get("/exportar/excel")
def exportar_excel(
    fecha_inicio: Optional[date] = Query(None, description="Fecha inicio"),
    fecha_fin: Optional[date] = Query(None, description="Fecha fin"),
    anio: Optional[int] = Query(None, description="Filtrar por año"),
    mes: Optional[int] = Query(None, ge=1, le=12, description="Filtrar por mes"),
    db: Session = Depends(get_db),
):
    """
    Exportar ingresos y gastos a un archivo Excel.
    Se puede filtrar por rango de fechas, año o mes.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ── Construir filtros ──
    filtros_ing = []
    filtros_gas = []

    if fecha_inicio:
        filtros_ing.append(Income.fecha >= fecha_inicio)
        filtros_gas.append(Expense.fecha >= fecha_inicio)
    if fecha_fin:
        filtros_ing.append(Income.fecha <= fecha_fin)
        filtros_gas.append(Expense.fecha <= fecha_fin)
    if anio:
        filtros_ing.append(extract("year", Income.fecha) == anio)
        filtros_gas.append(extract("year", Expense.fecha) == anio)
    if mes:
        filtros_ing.append(extract("month", Income.fecha) == mes)
        filtros_gas.append(extract("month", Expense.fecha) == mes)

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    income_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    expense_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    summary_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── Hoja de Ingresos ──
    ws_ing = wb.active
    ws_ing.title = "Ingresos"
    ws_ing.append(["ID", "Fecha", "Descripción", "Monto ($)"])
    for col in range(1, 5):
        cell = ws_ing.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    ingresos = (
        db.query(Income)
        .filter(*filtros_ing)
        .order_by(Income.fecha.desc())
        .all()
    )
    total_ingresos = 0.0
    for i, ing in enumerate(ingresos, start=2):
        ws_ing.append([ing.id, str(ing.fecha), ing.descripcion, ing.monto])
        total_ingresos += ing.monto
        for col in range(1, 5):
            cell = ws_ing.cell(row=i, column=col)
            cell.border = thin_border
            if col == 4:
                cell.number_format = "#,##0.00"

    fila_total = len(ingresos) + 2
    ws_ing.cell(row=fila_total, column=3, value="TOTAL INGRESOS").font = Font(bold=True)
    ws_ing.cell(row=fila_total, column=4, value=round(total_ingresos, 2)).font = Font(bold=True)
    ws_ing.cell(row=fila_total, column=3).fill = income_fill
    ws_ing.cell(row=fila_total, column=4).fill = income_fill

    for col in [ws_ing.column_dimensions["A"], ws_ing.column_dimensions["B"],
                ws_ing.column_dimensions["C"], ws_ing.column_dimensions["D"]]:
        col.width = 20

    # ── Hoja de Gastos ──
    ws_gas = wb.create_sheet("Gastos")
    ws_gas.append(["ID", "Fecha", "Descripción", "Categoría", "Monto ($)"])
    for col in range(1, 6):
        cell = ws_gas.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    gastos = (
        db.query(Expense)
        .filter(*filtros_gas)
        .order_by(Expense.fecha.desc())
        .all()
    )
    total_gastos = 0.0
    for i, gas in enumerate(gastos, start=2):
        cat_display = gas.categoria.value if hasattr(gas.categoria, "value") else gas.categoria
        ws_gas.append([gas.id, str(gas.fecha), gas.descripcion, cat_display, gas.monto])
        total_gastos += gas.monto
        for col in range(1, 6):
            cell = ws_gas.cell(row=i, column=col)
            cell.border = thin_border
            if col == 5:
                cell.number_format = "#,##0.00"

    fila_total_g = len(gastos) + 2
    ws_gas.cell(row=fila_total_g, column=4, value="TOTAL GASTOS").font = Font(bold=True)
    ws_gas.cell(row=fila_total_g, column=5, value=round(total_gastos, 2)).font = Font(bold=True)
    ws_gas.cell(row=fila_total_g, column=4).fill = expense_fill
    ws_gas.cell(row=fila_total_g, column=5).fill = expense_fill

    for col_letter in ["A", "B", "C", "D", "E"]:
        ws_gas.column_dimensions[col_letter].width = 22

    # ── Hoja Resumen ──
    ws_resumen = wb.create_sheet("Resumen")
    ws_resumen.append(["Concepto", "Monto ($)"])
    for col in range(1, 3):
        cell = ws_resumen.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    ws_resumen.append(["Total Ingresos", round(total_ingresos, 2)])
    ws_resumen.append(["Total Gastos", round(total_gastos, 2)])
    balance = round(total_ingresos - total_gastos, 2)
    ws_resumen.append(["Balance (Ingresos - Gastos)", balance])

    for row in range(2, 5):
        for col in range(1, 3):
            cell = ws_resumen.cell(row=row, column=col)
            cell.border = thin_border
            if col == 2:
                cell.number_format = "#,##0.00"

    ws_resumen.cell(row=4, column=1).fill = summary_fill
    ws_resumen.cell(row=4, column=2).fill = summary_fill
    ws_resumen.cell(row=4, column=1).font = Font(bold=True, size=13)
    ws_resumen.cell(row=4, column=2).font = Font(bold=True, size=13)

    # Gastos por categoría
    ws_resumen.append([])
    ws_resumen.append(["Gastos por Categoría", ""])
    ws_resumen.cell(row=6, column=1).font = Font(bold=True, size=12)

    cats = _gastos_por_categoria(db, filtros_gas)
    for cat in cats:
        ws_resumen.append([cat.categoria, cat.total])

    ws_resumen.column_dimensions["A"].width = 35
    ws_resumen.column_dimensions["B"].width = 20

    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = "reporte_contable"
    if anio:
        nombre_archivo += f"_{anio}"
    if mes:
        nombre_archivo += f"_{mes:02d}"
    if fecha_inicio and fecha_fin:
        nombre_archivo += f"_{fecha_inicio}_{fecha_fin}"
    nombre_archivo += ".xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )
